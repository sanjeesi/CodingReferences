import re
import openpyxl

emailTo = 'sanjeev.singh@amdocs.com'
emailFrom = 'qwerty@234'
subject = 'To do'
emailBody = '''
Hi,

As discussed with Mr. X the development for the CR123450 is complete.
We need you to do the following:
Task1: Test the new implementation XXXXXX in CR123450.
Task2: Send the testing report to Mr. Ed.
'''

def identifyTask(emailBody):
    taskRegex = re.compile(r'(?<=Task[ \d]:)[\. A-Za-z0-9]+')
    mo = taskRegex.findall(emailBody)
    print(mo)
    print('Assigned To: ' + emailTo)
    print('Allocated by: ' + emailFrom)

def writeOnExcel(time, sender, to, subject, body):
    loc = r"C:\Users\sanjeesi\Desktop\Workspace02\ProjectAviskar\TaskSheet.xlsx"
    wb = openpyxl.load_workbook(loc)
    sheet_wrt = wb['abc']
    c1 = sheet_wrt.cell(row = 2, column = 1)
    c1.value = time
    c2 = sheet_wrt.cell(row = 2, column = 2)
    c2.value = sender
    c3 = sheet_wrt.cell(row = 2, column = 3)
    c3.value = to
    c4 = sheet_wrt.cell(row = 2, column = 4)
    c4.value = subject
    c5 = sheet_wrt.cell(row = 2, column = 5)
    c5.value = body
    wb.save(loc)

# identifyTask(emailBody)
writeOnExcel(12, emailFrom, emailTo, subject, emailBody)