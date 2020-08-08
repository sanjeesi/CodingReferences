import xlrd
import openpyxl
from datetime import datetime
import requests

MAX_ROWS = 151

def callApi(siteAddress):
    API_ENDPOINT = 'https://maps.googleapis.com/maps/api/geocode/json'
    PARAMS = {'address':siteAddress, 'key':'YOUR_KEY'}
    proxy = {
                'http':'http://genproxy.ABC.com:8080',
                'https':'http://genproxy.ABC.com:8080'
            }
    res = requests.get(url = API_ENDPOINT, params = PARAMS, proxies = proxy)
    data = res.json()
    res.close()
    # print(longitude)
    return data

def readExcel(sheet, row):
    siteAddress = sheet.cell_value(row, 0)
    return siteAddress

def writeExcel(sheet, row1, data):
    try:
        latitude = data['results'][0]['geometry']['location']['lat']
        longitude = data['results'][0]['geometry']['location']['lng']
        c1 = sheet.cell(row = row1+1, column = 2)
        c1.value = latitude
        c2 = sheet.cell(row = row1+1, column = 3)
        c2.value = longitude
    except Exception as e:
        print(e)

loc = (r"sample_input.xlsx")
wb = xlrd.open_workbook(loc)
sheet_rd = wb.sheet_by_index(0)
wb1 = openpyxl.load_workbook(loc)
sheet_wrt = wb1.worksheets[0]
objectList = []
for index in range(1, MAX_ROWS):
    siteAddress = readExcel(sheet_rd, index)
    print(siteAddress)
    try:
        data = callApi(siteAddress)
    except Exception as e:
        print(e)
    writeExcel(sheet_wrt, index, data)
wb1.save(loc)