# import ctypes
from datetime import datetime
import pythoncom
import win32com.client
import openpyxl
import psycopg2
import requests

dataList=[]

def callApi(data):
    API_ENDPOINT = 'http://'
    res = requests.post(url = API_ENDPOINT, json = data)
    print(res)


def writeOnSqlTable(subject, sender, recipients, time, body):
    try:
        #Establishing the connection
        conn = psycopg2.connect(database="postgres", user='postgres', password='PostgreSql', host='127.0.0.1', port= '5432')
        #Setting auto commit false
        conn.autocommit = True

        #Creating a cursor object using the cursor() method
        cursor = conn.cursor()

        cursor.execute("select max(id+1) from task;")
        value = (cursor.fetchone())
        id = value[0]
        if id is None:
            id = 1
        # Preparing SQL queries to INSERT a record into the database.
        cursor.execute('''INSERT INTO task(id, Subject, Sender, Recipients, Time, Body)
        VALUES (%s, %s, %s, %s, %s, %s)''', (id, subject, sender, recipients, time, body))

        conn.commit()
        print('record inserted')
        conn.close()
    except Exception as e:
        print(e)

def writeOnExcel(time, sender, to, subject, body, row = 5):
    loc = r"C:\Workspace02\ProjectAviskar\TaskSheet.xlsx"
    wb = openpyxl.load_workbook(loc)
    sheet_wrt = wb['abc']
    c1 = sheet_wrt.cell(row = row, column = 1)
    c1.value = str(time)
    c2 = sheet_wrt.cell(row = row, column = 2)
    c2.value = sender
    c3 = sheet_wrt.cell(row = row, column = 3)
    c3.value = to
    c4 = sheet_wrt.cell(row = row, column = 4)
    c4.value = subject
    c5 = sheet_wrt.cell(row = row, column = 5)
    c5.value = body
    wb.save(loc)

class Handler_Class(object):

    def OnNewMailEx(self, receivedItemsIDs):
        for ID in receivedItemsIDs.split(","):
            mail = outlook.Session.GetItemFromID(ID)
            subject = mail.Subject
            body = mail.Body
            sender = mail.Sender.GetExchangeUser().PrimarySmtpAddress
            receiveTime = mail.ReceivedTime
            to = mail.Recipients
            objectList = []
            for recipient in to:
                recipientAddress = recipient.AddressEntry.GetExchangeUser().PrimarySmtpAddress
                uniqueId = str(int(datetime.now().timestamp() * 1000))
                # uniqueId = str(uniqueId)
                objectData = {
                    "uniqueTaskId": uniqueId,
                    "ownerName": recipientAddress,
                    "taskDetails": body,
                    "assignedBy": sender,
                    "priority": "Low",
                    "createdDate": receiveTime.strftime('%Y-%m-%dT%H:%M:%S'),
                    "updateDate": datetime.now().strftime('%Y-%m-%dT%H:%M:%S'),
                    "active": True
                }
                objectList.append(objectData)
            data = {
                "sourceSystem": "email",
                "ownerTaskModelList": objectList
            }
            try:
                print(data)
                callApi(data)

                # writeOnSqlTable(subject, sender, to, str(receiveTime), body)

                # print('Writing on Excel')
                # writeOnExcel(str(receiveTime),sender,to,subject,body,4)

                print('E-mail task created successfully')
            except Exception as e:
                print(e)

# writeOnExcel(12, 'ab', 'to', 'sub', 'body', 4)

print("Listening on Outlook Events")
outlook = win32com.client.DispatchWithEvents("Outlook.Application", Handler_Class)

# and then an infinit loop that waits from events.
pythoncom.PumpMessages()


#
#
# outlook = outlook = win32com.client.Dispatch("Outlook.Application").GetNameSpace("MAPI")
# inbox = outlook.GetDefaultFolder(6)  # Inbox default index value is 6\
#
# message = inbox.Items
# messages = message.GetLast()
# message.Sort("[ReceivedTime]", True)
# messages = message.Restrict("[ReceivedTime] >= '" + datetime.today.strftime('%m/%d/%Y %H:%M %p') + "'")
# while(True):
#
#         subject = messages.Subject
#         body = messages.body
#         date = messages.senton.date()
#         recievedTime=  messages.ReceivedTime
#         sender = messages.Sender
#         cc = messages.Cc
#         recipient = messages.To
#         attachments = messages.Attachments
#         if(subject.startswith('RE:')):
#                 print('NO!')
#
#         else:
#
#                 print('My time', recievedTime)
#                 print(subject)
#                 print(body)
#                 print(recipient)
#                 print(cc)
#                 print(attachments.count)
#                 print(date)
#
#         time.sleep(5)
#



